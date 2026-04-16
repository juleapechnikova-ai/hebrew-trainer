#!/usr/bin/env python3
"""Импорт уроков 1–19 из Excel в lessons.json (объединение с существующими 20–28)."""
import json
import re
import sys
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
LESSONS_JSON = ROOT / "src" / "data" / "lessons.json"
XLSX_DEFAULT = (
    Path(__file__).resolve().parents[2] / "уроки 1-19.xlsx"
)


def parse_lesson_id(cell) -> Optional[int]:
    if cell is None:
        return None
    s = str(cell).strip()
    if not s:
        return None
    m = re.match(r"^(\d+)", s)
    if m:
        return int(m.group(1))
    return None


def cell_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def import_xlsx(path: Path) -> Dict[str, List[dict]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    by_lesson: dict[int, list] = {i: [] for i in range(1, 20)}

    for row in ws.iter_rows(min_row=2, values_only=True):
        ru, he = row[1], row[2]
        ru_s, he_s = cell_str(ru), cell_str(he)
        if not ru_s or not he_s:
            continue
        lid = parse_lesson_id(row[7])
        if lid is None or lid < 1 or lid > 19:
            continue

        plural = cell_str(row[3])
        pos = cell_str(row[5])
        gender = cell_str(row[6])
        notes = cell_str(row[8])

        item = {
            "ru": ru_s,
            "he": he_s,
            "lessonId": lid,
            "pos": pos or "слово",
            "gender": gender,
            "notes": notes,
        }
        if plural:
            item["plural"] = plural

        by_lesson[lid].append(item)

    wb.close()
    return {str(k): v for k, v in by_lesson.items()}


def main():
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else XLSX_DEFAULT
    if not xlsx.is_file():
        print(f"Файл не найден: {xlsx}", file=sys.stderr)
        sys.exit(1)

    new_chunks = import_xlsx(xlsx)
    with open(LESSONS_JSON, encoding="utf-8") as f:
        data = json.load(f)

    for key, items in new_chunks.items():
        data[key] = items

    keys = sorted(int(k) for k in data.keys())
    with open(LESSONS_JSON, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")

    print(f"OK: {LESSONS_JSON}")
    print("Ключи уроков:", keys)
    total = sum(len(data[str(k)]) for k in range(1, 20))
    print(f"Слов в уроках 1–19: {total}")


if __name__ == "__main__":
    main()
